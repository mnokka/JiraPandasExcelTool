# skipped Pandas and using tradional excel reading lib: import openpyxl 
#
# Jira Exp/Imp used to export excel with all field information
# This utility ut use excel info to move issue data (fields) to another project in another Jira
# 


from jira import JIRA
from datetime import datetime
import logging as log
#import pandas 
import argparse
import getpass
import time
import sys, logging
from author import Authenticate  # no need to use as external command
from author import DoJIRAStuff
import openpyxl 
from collections import defaultdict
import re
from pprint import pprint

start = time.clock()
__version__ = u"0.9.RISKS" 



###################################################################
# should pass via parameters
# CODE CONFIGURATIONS
#####################################################################

# development vs production Jira
#ENV="DEV"
ENV="PROD"

#risk vs mitigation risk project operations
#TYPE="MITI"
TYPE="RISK"

#project type ship vs. finance
CAT="SHIP"
#CAT="FIN"

# do only one operation for testing purposes
#ONCE="NO"
ONCE="YES"

###########################################################################


logging.basicConfig(level=logging.DEBUG) # IF calling from Groovy, this must be set logging level DEBUG in Groovy side order these to be written out


def main():

    
    JIRASERVICE=u""
    JIRAPROJECT=u""
    PSWD=u''
    USER=u''
  
    logging.debug (u"--Python starting checking Jira issues for attachemnt adding --") 

 
    parser = argparse.ArgumentParser(usage="""
    {1}    Version:{0}     -  mika.nokka1@gmail.com 
    
    USAGE:
    -filepath  | -p <Path to Excel file directory>
    -filename   | -n <Excel filename>
    TODO ADD DESCR
    PLEASE SEE THE CODE

    """.format(__version__,sys.argv[0]))

    #parser.add_argument('-f','--filepath', help='<Path to attachment directory>')
    parser.add_argument('-q','--excelfilepath', help='<Path to excel directory>')
    parser.add_argument('-n','--filename', help='<Issues excel filename>')
    
    parser.add_argument('-v','--version', help='<Version>', action='store_true')
    
    parser.add_argument('-w','--password', help='<JIRA password>')
    parser.add_argument('-u','--user', help='<JIRA username>')
    parser.add_argument('-s','--service', help='<JIRA service, like https://my.jira.com>')
    parser.add_argument('-l','--links', help='<Target Jira project ID to which these issues to be linked, if link info (linked issue summary) excel>') #add issue links to generated issues (target "into" linked issues must be allready in target jira)
    parser.add_argument('-p','--project', help='<Target JIRA project ID to be used>')
    #parser.add_argument('-z','--rename', help='<rename files>') #adhoc operation activation
    #parser.add_argument('-x','--ascii', help='<ascii file names>') #adhoc operation activation
        
    args = parser.parse_args()
    
    if args.version:
        print 'Tool version: %s'  % __version__
        sys.exit(2)    
           
    #filepath = args.filepath or ''
    excelfilepath = args.excelfilepath or ''
    filename = args.filename or ''
    
    JIRASERVICE = args.service or ''
    JIRAPROJECT = args.project or ''
    PSWD= args.password or ''
    USER= args.user or ''
    LINKS=args.links or ''
    #RENAME= args.rename or ''
    #ASCII=args.ascii or ''
    
    # quick old-school way to check needed parameters
    if (JIRASERVICE=='' or PSWD=='' or USER==''  or excelfilepath=='' or JIRAPROJECT=='' or filename==''):
        parser.print_help()
        print "args: {0}".format(args)
        sys.exit(2)

    
    
    Authenticate(JIRASERVICE,PSWD,USER)
    jira=DoJIRAStuff(USER,PSWD,JIRASERVICE)
    
    excel=excelfilepath+"/"+filename
    logging.debug ("Excel file:{0}".format(excel))

    Issues=defaultdict(dict) 
    MainSheet="Sheet0" 
    wb= openpyxl.load_workbook(excel)
    #types=type(wb)
    #logging.debug ("Type:{0}".format(types))
    #sheets=wb.get_sheet_names()
    #logging.debug ("Sheets:{0}".format(sheets))
    CurrentSheet=wb[MainSheet] 
    
    
    ########################################
    #CONFIGURATIONS AND EXCEL COLUMN MAPPINGS, both main and subtask excel
    DATASTARTSROW=4 # data section starting line 
    A=1 #issuetype
    E=5 #SUMMARY
    F=6 #priority
    H=8 #Status   
    O=15   #system number
    Q=17 #Assignee
    S=19 #Disciopline(F)
    K=11 #duedate
    
    T=20 #Probability
    U=21 #HSE Impact
    V=22 #Schedule Impact 
    W=23 #Quality Impact
    
    Z=26 #Risk Cost
    AK=37 #Linked Issues
    AM=39 ##Disciopline(RM)
    AN=40 #Description
    AB=28 #Mitigation Costs (Keur)
    
 
    print "=====>    Internal configuration:{0} , {1} , {2}".format(ENV, TYPE, CAT)
 
    ##############################################################################################
    #Go through main excel sheet for main issue keys (and contents findings)
    # Create dictionary structure (remarks)
    # NOTE: Uses hardcoded sheet/column values
    # NOTE: As this handles first sheet, using used row/cell reading (buggy, works only for first sheet) 
    #
    i=DATASTARTSROW # brute force row indexing
    for row in CurrentSheet[('C{}:C{}'.format(DATASTARTSROW,CurrentSheet.max_row))]:  # go trough all column C  rows (issue key when imp exp eported)
        for mycell in row:
            KEY=mycell.value
            logging.debug("ROW:{0} Original ID:{1}".format(i,mycell.value))
            Issues[KEY]={} # add to dictionary as master key (KEY)
            
            #Just hardocode operations, POC is one off
            #LINKED_ISSUES=(CurrentSheet.cell(row=i, column=K).value) #NOTE THIS APPROACH GOES ALWAYS TO THE FIRST SHEET
            #logging.debug("Attachment:{0}".format((CurrentSheet.cell(row=i, column=K).value))) # for the same row, show also column K (LINKED_ISSUES) values
            #Issues[KEY]["LINKED_ISSUES"] = LINKED_ISSUES
            
            SUMMARY=(CurrentSheet.cell(row=i, column=E).value)
            if not SUMMARY:
                SUMMARY="Summary for this task has not been defined"
            Issues[KEY]["SUMMARY"] = SUMMARY
            
            ISSUE_TYPE=(CurrentSheet.cell(row=i, column=A).value)
            Issues[KEY]["ISSUE_TYPE"] = ISSUE_TYPE
            
            STATUS=(CurrentSheet.cell(row=i, column=E).value)
            Issues[KEY]["SUMMARY"] = SUMMARY
            
            PRIORITY=(CurrentSheet.cell(row=i, column=F).value)
            Issues[KEY]["PRIORITY"] = PRIORITY
           
            
            STATUS=(CurrentSheet.cell(row=i, column=H).value)
            Issues[KEY]["STATUS"] = STATUS
            
            
            ASSIGNEE=(CurrentSheet.cell(row=i, column=Q).value)
            Issues[KEY]["ASSIGNEE"] = ASSIGNEE
            
            DisciplineF=(CurrentSheet.cell(row=i, column=S).value)
            Issues[KEY]["DisciplineF"] = DisciplineF
            
            DisciplineRM=(CurrentSheet.cell(row=i, column=AM).value)
            Issues[KEY]["DisciplineRM"] = DisciplineRM
            
            DESCRIPTION=(CurrentSheet.cell(row=i, column=AN).value)
            Issues[KEY]["DESCRIPTION"] = DESCRIPTION
            
            PROBABILITY=(CurrentSheet.cell(row=i, column=T).value)
            Issues[KEY]["PROBABILITY"] = PROBABILITY
            
            HSEImpact=(CurrentSheet.cell(row=i, column=U).value)
            Issues[KEY]["HSEImpact"] = HSEImpact
            
              
            #RESPHONE=(CurrentSheet.cell(row=i, column=U).value)
            #Issues[KEY]["RESPHONE"] = RESPHONE
            
            SheduleImpact=(CurrentSheet.cell(row=i, column=V).value)
            Issues[KEY]["SheduleImpact"] = SheduleImpact
            
            QualityImpact=(CurrentSheet.cell(row=i, column=W).value)
            Issues[KEY]["QualityImpact"] = QualityImpact
            
            RiskCost=(CurrentSheet.cell(row=i, column=Z).value)
            Issues[KEY]["RiskCost"] = RiskCost
        
            MitigationCostsKeur=(CurrentSheet.cell(row=i, column=AB).value)
            Issues[KEY]["MitigationCostsKeur"] = MitigationCostsKeur
        
                
            LinkedIssues=(CurrentSheet.cell(row=i, column=AK).value)
            Issues[KEY]["LinkedIssues"] = LinkedIssues
            
            
            DueDate=(CurrentSheet.cell(row=i, column=K).value)
            Issues[KEY]["DueDate"] = str(DueDate)  # othewise datetime object
            
            SystemNumber=(CurrentSheet.cell(row=i, column=O).value)
            Issues[KEY]["SystemNumber"] = SystemNumber

            logging.debug("---------------------------------------------------")
            i=i+1
    
    #print Issues.items() 
  
    #print "priority after all settings:{0}".format(PRIORITY)  
    for key, value in Issues.iteritems() :
        print "\n\n++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        print "ORIGINAL ISSUE KEY: {0}".format(key)
        #print key, value
        
        # check linked issues values form each issue found from excel
        one=Issues.get(key)
        for key, value in one.iteritems() :
            print "************************************************************************"
            if (value==None):
                if (isinstance(value, long)):
                    castedValue=0
                    print "NONE NUMBER"
                else:
                    castedValue=0
                    print "NONE" 
            else:
                if (isinstance(value, long)): # is it number??
                    castedValue=value # numbers dont need utf-8 endocing
                else:
                    castedValue=value.encode('utf-8') 
                  
            
            #print "value:{0}".format(value)
            print "{0} {1}".format(key,castedValue)
            #print "priority in loop:{0}".format(PRIORITY)  
            
            if (LINKS and key=="LinkedIssues"): #-l parameter to do links operations to given target project
                #print "Linked issues column found"
                
                TOLINKLIST=[] # all to be linked issues for this issue stored here for processing
                print "Linking active: Linking target project: {0}".format(LINKS)
                if (value==None): #no linked items case
                    value2="NONE"
                    print "No linked issues found"
                else:
                    value2=value
                onelink=value2.split(':')
                for item in onelink :
                    #print "value:{0}".format(item.encode('utf-8'))
                    regex = r"(.*)(')(.*)(')"   #TT1400-39 'Logistic plan to do' (Risk Mitigation)
                    match = re.search(regex, item)
                
                    if (match):
                        hit=match.group(3)
                        #print "-----------------------------------------------------------"
                        print "Linked issue Summmary ==>  {0}".format(hit.encode('utf-8'))
                        #print "-----------------------------------------------------------"
                        
                        #project = "Risk Mitigation Panel Line"  and summary ~ "Kuitulaser hankinta ja hitsauslaboratorion hankinta"
                        #issue_list = jira.search_issues("Project = {0} and Summary ~ {1}".format(LINKS,hit))
                        
                        jql_query="Project = \'{0}\' and Summary ~ \'{1}\'".format(LINKS,hit.encode('utf-8'))
                        #print "Query:{0}".format(jql_query)
                        
                        issue_list=jira.search_issues(jql_query)
                        
                        if len(issue_list) >= 1:
                            for issue in issue_list:
                                #logging.debug("One issue returned for query")
                                logging.debug("ISSUE TO BE LINKED ==> {0}".format(issue))
                                LINKEDISSUE=issue
                                TOLINKLIST.append(issue)
                        #elif len(issue_list) > 1:
                        #    logging.debug("ERROR ==> More than 1 issue was returned by JQL query")
                        #    LINKEDISSUE="EMPTY"
                        else:
                            logging.debug("==> No issue(s) returned by JQL query")
                            #LINKEDISSUE="EMPTY"
            #else:
            #    LINKEDISSUE="EMPTY"               
                        
                
            if (key=="ASSIGNEE"):
                #print "Assignee column found"
                
                if (value==None): #no linked items case
                    value2="NONE"
                    print "No assignee found"
                    USERNAME_ASSIGNEE="-1"
                else:
                    value2=value
                    regex = r"(.*)(\()(.*)(\))"   #Korpela, Matias (korpma11)
                    match = re.search(regex, value2)
                
                    if (match):
                        USERNAME_ASSIGNEE=match.group(3).encode('utf-8')
                        #print "-----------------------------------------------------------"
                        print "Assignee username ==>  {0}".format(USERNAME_ASSIGNEE)
                        #print "-----------------------------------------------------------"
                    else:
                        USERNAME_ASSIGNEE="-1"
                        
                
            if (key=="MitigationCostsKeur"):
                #print "Mitigation cost column found"
                MitigationCostsKeur=castedValue
                       
            if (key=="STATUS"):
                #print "STATUS cost column found: {0}".format(value)
                if (ENV =="DEV" and TYPE=="MITI"):
                    if (value=="To Do"):
                        print "Dev: To Do found, doing nothing"
                        NEWSTATUS="To Do"
                    else:
                        NEWSTATUS=value  
                        print "Dev: new status set:{0}".format(NEWSTATUS)
                
                elif (ENV =="DEV" and TYPE=="RISK"):
                    if (value=="Proposed"):
                        print "Dev: TProposed found, doing nothing"
                        NEWSTATUS="Proposed"
                    else:
                        NEWSTATUS=value  
                        print "Dev: new status set:{0}".format(NEWSTATUS)
                
                #duplicating intentionally       
                elif (ENV =="PROD" and TYPE=="MITI"):
                    if (value=="To Do"):
                        print "Prod: To Do found, doing nothing"
                        NEWSTATUS="To Do"
                    else:
                        NEWSTATUS=value  
                        print "Prod: new status set:{0}".format(NEWSTATUS)
                
                elif (ENV =="PROD" and TYPE=="RISK"):
                    if (value=="Proposed"):
                        print "Prod: TProposed found, doing nothing"
                        NEWSTATUS="Proposed"
                    else:
                            NEWSTATUS=value  
                            print "Prod: new status set:{0}".format(NEWSTATUS)
            
                else:
                    print "ERROR: NO STATUSES"
                    sys.exit(5)        
                    #TODO STATUSES"    
                
                    
                    
                    
            
            if (key=="SUMMARY"):
                SUMMARY=castedValue  
            
            if (key=="ISSUE_TYPE"):
                ISSUE_TYPE=castedValue 
                
            if (key=="PRIORITY"):
                PRIORITY=castedValue    
            
            if (key=="DESCRIPTION"):
                DESCRIPTION=castedValue 
                
            if (key=="DisciplineRM"):
                if(castedValue==0):
                    DisciplineRM="-1"  
                else:
                    DisciplineRM=castedValue    
            
            if (key=="DisciplineF"):
                if(castedValue==0):
                    DisciplineF="-1"      
                else:
                    DisciplineF=castedValue  
            
            

            if (CAT=="SHIP"):
                DISCIPLINE=DisciplineRM
            elif (CAT=="FIN"):
                DISCIPLINE=DisciplineF
            
             
            
            if (key=="SystemNumber"):
                SystemNumber=castedValue   
                
            if (key=="RiskCost"):
                RiskCost=castedValue   
                   
            if (key=="HSEImpact"):
                HSEImpact=castedValue   
            
            if (key=="PROBABILITY"):
                PROBABILITY=castedValue  
            
            if (key=="QualityImpact"):
                QualityImpact=castedValue  
                
            if (key=="SheduleImpact"):
                SheduleImpact=castedValue     
     
                
            if (key=="DueDate"):
                DueDate=castedValue
                if (not(DueDate) or DueDate=="None"):
                    DueDate="0"
                else: #2019-03-01 00:00:00    (\d\d\d\d-\d\d-\d\d)(.*)
                    regex = r"(\d\d\d\d-\d\d-\d\d)(.*)"   #TT1400-39 'Logistic plan to do' (Risk Mitigation)
                    match = re.search(regex, DueDate)
                
                    if (match):
                        hit=match.group(1)
                        DueDate=hit
                    else:
                        print "Error for DueDate parsing"
                    
                     
            
                        
        # just 2 funcitons for 2 projectypes, this is just a tool
        if (TYPE=="MITI"):
            CreateMitigationIssue(jira,JIRAPROJECT,SUMMARY,ISSUE_TYPE,PRIORITY,STATUS,USERNAME_ASSIGNEE,DESCRIPTION,MitigationCostsKeur,NEWSTATUS,ENV,DISCIPLINE,CAT,DueDate)
        elif (TYPE=="RISK"):
            print "Calling Discipline:{0}".format(DISCIPLINE)
            CreateRiskIssue(jira,JIRAPROJECT,SUMMARY,ISSUE_TYPE,PRIORITY,STATUS,USERNAME_ASSIGNEE,DESCRIPTION,MitigationCostsKeur,NEWSTATUS,ENV,DISCIPLINE,TYPE,RiskCost,CAT,TOLINKLIST,LINKS,DueDate,SystemNumber,HSEImpact,PROBABILITY,QualityImpact,SheduleImpact)
        else:
            print "Lost in translation. Cant do want I should do"
                
        time.sleep(0.7) # prevent jira crashing for script attack
        if (ONCE=="YES"):
            print "ONCE testing mode ,stopping now"
            sys.exit(5) #testing do only once
        print "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        #now excel has been prosessed
        
    end = time.clock()
    totaltime=end-start
    print "Time taken:{0} seconds".format(totaltime)
    print "*************************************************************************"
    sys.exit(0)
    




    
def CreateMitigationIssue(jira,JIRAPROJECT,SUMMARY,ISSUE_TYPE,PRIORITY,STATUS,USERNAME_ASSIGNEE,DESCRIPTION,MitigationCostsKeur,NEWSTATUS,ENV,DISCIPLINE,CAT,DueDate):
    
    
    TRANSIT="None"
    jiraobj=jira
    project=JIRAPROJECT
    TASKTYPE="Task" #hardcoded

    print "Creating mitigation issue for JIRA project: {0}".format(project)
    
    issue_dict = {
    'project': {'key': JIRAPROJECT},
    'summary': str(SUMMARY),
    'description': str(DESCRIPTION),
    'issuetype': {'name': TASKTYPE},
    'priority': {'name': str(PRIORITY) }, 
    'assignee': {'name':USERNAME_ASSIGNEE},
    'customfield_14302' if (ENV =="DEV") else 'customfield_14216' : int(MitigationCostsKeur), # MitigationCostsKeur dev: 14302  prod: 14216

    }

    try:
        new_issue = jiraobj.create_issue(fields=issue_dict)
        print "===> Issue created OK:{0}".format(new_issue)
        if (NEWSTATUS != "To Do"): # status after cretion
            
            #map state to transit for Mitigation issues
            if (NEWSTATUS=="In Progress"):
                TRANSIT="Start Progress"
            if (NEWSTATUS=="Done"):
                TRANSIT="Done"
            
            
            print "Newstatus will be:{0}".format(NEWSTATUS)
            print "===> Executing transit:{0}".format(TRANSIT)
            jiraobj.transition_issue(new_issue, transition=TRANSIT)  # trantsit to state where it was in excel 
        else:
            print "Initial status found: {0}, nothing done".format(NEWSTATUS)
            
        if not (DueDate=="0"):
            new_issue.update(duedate=DueDate)
        
        # hack as Jira started to assign the user of the script during transit phase        
        jiraobj.assign_issue(new_issue, USERNAME_ASSIGNEE)    
   
        
    except Exception,e:
        print("Failed to create JIRA object or transit problem, error: %s" % e)
        sys.exit(1)
    return new_issue    
    
     
def CreateRiskIssue(jira,JIRAPROJECT,SUMMARY,ISSUE_TYPE,PRIORITY,STATUS,USERNAME_ASSIGNEE,DESCRIPTION,MitigationCostsKeur,NEWSTATUS,ENV,DISCIPLINE,TYPE,RiskCost,CAT,TOLINKLIST,LINKS,DueDate,SystemNumber,HSEImpact,PROBABILITY,QualityImpact,SheduleImpact):
    
    print "=====>    Internal configuration:{0} , {1} , {2}".format(ENV, TYPE, CAT)
    print "Discipline:{0} ".format(DISCIPLINE)
    
    TRANSIT="NA"
    jiraobj=jira
    project=JIRAPROJECT
    TASKTYPE="Task" #hardcoded
    DISCIPLINEFIELD="None"

    print "Creating Risk issue for JIRA project: {0}".format(project)
    
    
    issue_dict = {
    'project': {'key': JIRAPROJECT},
    'summary': str(SUMMARY),
    'description': str(DESCRIPTION),
    'issuetype': {'name': TASKTYPE},
    'priority': {'name': str(PRIORITY) }, 
    #'resolution':{'id': '10100'},
    'assignee': {'name':USERNAME_ASSIGNEE}, 
    'customfield_14203' if (ENV =="DEV") else 'customfield_14208' : int(RiskCost),  # Risk Cost (Keur) dev: 14203  prod: 14208
    'customfield_14350' if (ENV =="DEV") else 'customfield_14212' : str(SystemNumber),  
    
    }

    try:
        new_issue = jiraobj.create_issue(fields=issue_dict)
        print "===> Issue created OK:{0}".format(new_issue)
        if (NEWSTATUS != "Proposed"): # status after cretion
            
            #map state to transit for Mitigation issues
            if (NEWSTATUS=="Threat"):
                TRANSIT="Threat"
            if (NEWSTATUS=="Realized"):
                TRANSIT="Realized"
            if (NEWSTATUS=="Eliminated"):
                TRANSIT="Eliminated"   
            if (NEWSTATUS=="No Action"):
                TRANSIT="No Action" # prod transt, dev transit was NoAction 
            
            print "Newstatus will be:{0}".format(NEWSTATUS)
            print "===> Executing transit:{0}".format(TRANSIT)
            jiraobj.transition_issue(new_issue, transition=TRANSIT)  # trantsit to state where it was in excel
        else:
            print "Initial status found: {0}, nothing done".format(NEWSTATUS)
            
        
        if not (DueDate=="0"):
            new_issue.update(notify=False,duedate=DueDate)
        
        
        
        
        #only quikc way set drop down menus, creation did not work as dictionary in use (should have used multiple dictionaries....)
        #print "Using dicipline:{0}".format(DISCIPLINE)
        if (ENV =="DEV" and CAT=="FIN"):
            DISCIPLINEFIELD="customfield_14223" # DisciplineF 
            new_issue.update(fields={DISCIPLINEFIELD: {'value': DISCIPLINE}})  #   DISCIPLIN 
        elif (ENV =="DEV" and CAT=="SHIP"):
            DISCIPLINEFIELD="customfield_14328" #  DisciplineRM
            new_issue.update(fields={DISCIPLINEFIELD: {'value' : DISCIPLINE}})  #   DISCIPLIN 
        elif (ENV =="PROD" and CAT=="FIN"):
            DISCIPLINEFIELD="customfield_14210" # DisciplineF 
            new_issue.update(fields={DISCIPLINEFIELD: {'value' : DISCIPLINE}})  #   DISCIPLIN 
        elif (ENV =="PROD" and CAT=="SHIP"): 
            DISCIPLINEFIELD="customfield_14209" #  DisciplineRM
            new_issue.update(notify=False,fields={DISCIPLINEFIELD: {'value' : DISCIPLINE}})  #   DISCIPLIN 
        else:
            print "ARGH ERRORS WTIH RISK DISCIPLINE FIELDS"    
            print "DISCIPLINE:{0}".format(DISCIPLINE)
            new_issue.update(fields={DISCIPLINEFIELD: {"id": "-1"}})  #   DISCIPLINE possible fails
        
        
        #these fields do have default value by Jira, mostly...
        if (ENV =="DEV"):
            HSEImpactFIELD="customfield_14224"
            if not(HSEImpact==0):
                HSEImpactFIELD="customfield_14224"  
                new_issue.update(fields={HSEImpactFIELD: {'value': HSEImpact}}) 
            else:
                new_issue.update(fields={HSEImpactFIELD: {"id": "-1"}}) 
            
            PROBABILITYFIELD="customfield_14225"
            if not(PROBABILITY==0):
                PROBABILITYFIELD="customfield_14225"  
                new_issue.update(fields={PROBABILITYFIELD: {'value' : PROBABILITY}})
            else:
                new_issue.update(fields={PROBABILITYFIELD: {"id": "-1"}}) 
            
            QualityImpactFIELD="customfield_14226"  
            if not(QualityImpact==0):
                QualityImpactFIELD="customfield_14226"  
                new_issue.update(fields={QualityImpactFIELD: {'value' : QualityImpact}})
            else:
                new_issue.update(fields={QualityImpactFIELD: {"id": "-1"}})
            
            SheduleImpactFIELD="customfield_14227" 
            if not(SheduleImpact==0):
                SheduleImpactFIELD="customfield_14227"  
                new_issue.update(fields={SheduleImpactFIELD: {'value' :   SheduleImpact}})
            else:
                new_issue.update(fields={SheduleImpactFIELD: {"id": "-1"}})
            
        
        #TODO FIX ME   *****************************************
        if (ENV =="PROD"):
            HSEImpactFIELD="customfield_14204"  
           
            if not(HSEImpact==0):
                HSEImpactFIELD="customfield_14204"  
                #new_issue.update(notify=False,fields={HSEImpactFIELD: {"id" : "26010" }}) # this works, ID from Jira UI (edit single selection ui), code uses meta info for this
                # goovy listener adds \'s for the string
                # EXCEL: <span class="amb-tooltip" style="display:inline-block;padding:.25rem .5rem;" title="Fatality(s) and/or permanent disability(s)">5</span><script>AJS.$("#customfield_14226-val .amb-tooltip").tooltip({gravity: 's'});</script>           
                # META:  <span class="amb-tooltip" style="display:inline-block;padding:.25rem .5rem;" title="Fatality(s) and/or permanent disability(s)">5</span><script>AJS.$("#customfield_14204-val .amb-tooltip").tooltip({gravity: \'s\'});</script>
                #REPLACED<span class="amb-tooltip" style="display:inline-block;padding:.25rem .5rem;" title="Fatality(s) and/or permanent disability(s)">5</span><script>AJS.$("#customfield_14226-val .amb-tooltip").tooltip({gravity: \'s\'});</script>
                HSEImpact2=HSEImpact.replace("'s'","\\'s\\'")
               
                regex = r"(.)(>\d<)(.*)"   # just get the >number< part to be used in in (whole string messes the operation
                match = re.search(regex, HSEImpact)
                
                if (match):
                    hit=match.group(2)
                    meta = jira.editmeta(new_issue)
                    #pprint(meta)     
                    field = None
                    for field in meta["fields"].values():
                        if field["name"] == "HSE Impact":
                            break
                    
                    value = [v for v in field["allowedValues"] if hit in v["value"]][0] # get id for excel based custom field string , use id to set the new value (string itself is too complicated and do not work)
                    new_issue.update(notify=False,fields={HSEImpactFIELD: {"id": value["id"]}}) # this works           
                else:
                    print "Error: can't decode excel based custom field value"     
            else:
                print "HSEImpact setting -1"
                new_issue.update(notify=False,fields={HSEImpactFIELD: {"id": "-1"}}) 
            
            
            
            
            
            
            PROBABILITYFIELD="customfield_14203" 
            if not(PROBABILITY==0):
                PROBABILITYFIELD="customfield_14203"  
                new_issue.update(notify=False,fields={PROBABILITYFIELD: {'value' : PROBABILITY}})
            else:
                new_issue.update(notify=False,fields={PROBABILITYFIELD: {"id": "-1"}})
        
        
            QualityImpactFIELD="customfield_14205"  
            if not(QualityImpact==0):
                QualityImpactFIELD="customfield_14205"  
                new_issue.update(notify=False,fields={QualityImpactFIELD: {'value' : QualityImpact}})
            else:
                new_issue.update(notify=False,fields={QualityImpactFIELD: {"id": "-1"}})
        
            SheduleImpactFIELD="customfield_14206"  
            if not(SheduleImpact==0):
                SheduleImpactFIELD="customfield_14206"  
                new_issue.update(notify=False,fields={SheduleImpactFIELD: {'value' :   SheduleImpact}})
            else:
                new_issue.update(notify=False,fields={SheduleImpactFIELD: {"id": "-1"}})
        

        
        
        
        
        #print "new issue: {0}   linked issue:{1}".format(new_issue,LINKEDISSUE)
        LENGHT=len(TOLINKLIST)
        print "List of linked ones, length:{0}".format(LENGHT)
        if (LINKS and TOLINKLIST): # link only if requested and there is something to link
            
            for LINKEDISSUE in TOLINKLIST:
                print "Linking requested, doing: new issue: {0} --> is mitigated by --->  linked issue:{1}".format(new_issue,LINKEDISSUE) # linktype hardcoded
                time.sleep(0.5)
                jiraobj.create_issue_link("is mitigated by",new_issue,LINKEDISSUE,None) # last is comment field, skipping now
        else:
            print "No linking requested nor no links for this issue, skipping"

        
        
        
    except Exception,e:
        #print("Failed to create JIRA object or transit problem, error: %s" % e)
        raise
        sys.exit(1)
    return new_issue   
    
    
if __name__ == '__main__':
    main()
    
    
    
    

    
    
    